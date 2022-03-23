from openpyxl import Workbook
from selenium import webdriver
import datetime
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

driver = webdriver.Chrome(executable_path= r"D:\Durai\Driver\chromedriver.exe")

date = datetime.date.today().strftime("%d-%m-%Y")
l = 2


wb = Workbook()
ws = wb.active
ws.title = "POORVIKA TV"


ws.cell(row=1, column=1).value = datetime.datetime.today()

driver.get(url="https://www.poorvika.com/television")
length = len(driver.find_elements(By.CLASS_NAME, "paginationid"))

print(length)
for r in range(1, length+1):
    print(r)
    driver.get(url="https://www.poorvika.com/television?&fmin=14999&fmax=68999&order=l-h&page=" + str(r))
    driver.implicitly_wait(1)
    # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    for cat in driver.find_elements(By.CLASS_NAME, "right-block"):
        name = cat.find_element(By.TAG_NAME, "h4")
        variant = cat.find_element(By.TAG_NAME, "h6")
        print(name.text,variant.text)
        ws.cell(row=l, column=1).value = name.text + variant.text
        price = cat.find_element(By.CLASS_NAME, "cat-price-new")
        print(price.text[1:])
        ws.cell(row=l, column=2).value = price.text[1:]
        wb.save(r"D:\Durai\Scraping\TV\Save Data\Poorvika File\Poorvika_TV "+date+".xlsx")
        l = l + 1
driver.quit()