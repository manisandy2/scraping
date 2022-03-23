from openpyxl import load_workbook,Workbook
from selenium import webdriver

from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
import datetime
import time
# 'cookies': 2, 'images': 2, 'javascript': 2, 'plugins': 2, 'popups': 2,
start_time = time.time()
options = webdriver.ChromeOptions()
prefs = {'profile.default_content_setting_values': {'geolocation': 2,
                            'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,
                            'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,
                            'media_stream_mic': 2, 'media_stream_camera': 2, 'protocol_handlers': 2,
                            'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                            'push_messaging': 2, 'ssl_cert_decisions': 2, 'metro_switch_to_desktop': 2,
                            'protected_media_identifier': 2, 'app_banner': 2, 'site_engagement': 2,
                            'durable_storage': 2}}
options.add_experimental_option('prefs', prefs)
# options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
# options.headless = True

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe",chrome_options=options)

date = datetime.date.today().strftime("%#d-%#m-%Y")
print(date)


wb = load_workbook(r"D:\Durai\Scraping\TV\Save Data\Save Urls\Scraping Urls "+date +".xlsx")
ws = wb.active
new_wb = Workbook()
new_ws = new_wb.active

new_ws.cell(row=1,column=1).value = "model Name"
new_ws.cell(row=1,column=2).value = "Poorvika price"
new_ws.cell(row=1,column=3).value = "Flipkart price"
new_ws.cell(row=1,column=4).value = "Amazon price"
new_ws.cell(row=1,column=5).value = "Croma price"
new_ws.cell(row=1,column=6).value = "vijay sale price"
new_ws.cell(row=1,column=7).value = "Reliance Digital price"

for r in range(2, ws.max_row+1):
    print(r)
    new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value
    new_ws.cell(row=r, column=2).value = ws.cell(row=r, column=2).value
    if ws.cell(row=r,column=3).value != "N/A":
# Flipkart
        print(ws.cell(row=r,column=3).value)
        try:
            driver.implicitly_wait(1)
            driver.get(url=ws.cell(row=r,column=3).value)
            new_ws.cell(row=r,column=3).value = driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:]
            print(driver.find_element(By.CLASS_NAME, "_30jeq3").text)
        except:
            pass
    #Amazon
    driver.implicitly_wait(3)
    if ws.cell(row=r, column=4).value != "N/A":
        print(ws.cell(row=r, column=4).value)
        try:
            driver.get(url=ws.cell(row=r, column=4).value)
            new_ws.cell(row=r, column=4).value = driver.find_element(By.CLASS_NAME, "a-text-price").text[1:]
            print("Amazon")
            print(driver.find_element(By.CLASS_NAME, "a-text-price").text[1:])

        except:
            pass

    # Croma
    driver.implicitly_wait(1)
    if ws.cell(row=r, column=5).value != "N/A":
        print(ws.cell(row=r, column=5).value)
        try:
            driver.get(url=ws.cell(row=r, column=5).value)
            for price in driver.find_elements(By.CLASS_NAME, "cp-price"):
                for price1 in price.find_elements(By.CLASS_NAME, 'new-price'):
                    for price2 in price1.find_elements(By.CLASS_NAME, 'amount'):
                        print(price2.text)
                        new_ws.cell(row=r, column=5).value = price2.text
        except:
            pass
    # vijay sale
    driver.implicitly_wait(1)
    if ws.cell(row=r, column=7).value != "N/A":
        print(ws.cell(row=r, column=7).value)
        driver.get(url=ws.cell(row=r, column=7).value)
        try:
            if driver.find_element(By.ID, "ContentPlaceHolder1_fillprice").text != None:
                # print("Nothing")
                try:
                    print(driver.find_element(By.XPATH,
                                              '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text)
                    price = driver.find_element(By.XPATH,
                                                '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text
                    new_ws.cell(row=r, column=6).value = price
                except:
                    print(driver.find_element(By.XPATH,
                                              '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text)
                    price = driver.find_element(By.XPATH,
                                                '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text
                    new_ws.cell(row=r, column=6).value = price
        except:
            pass
    # Reliance
    driver.implicitly_wait(1)
    if ws.cell(row=r, column=6).value != "N/A":
        print(ws.cell(row=r, column=6).value)
        try:
            driver.get(url=ws.cell(row=r, column=6).value)
            price = driver.find_element(By.CLASS_NAME, "pdp__offerPrice").text
            new_ws.cell(row=r, column=7).value = price[1:]
            print(price[1:])
        except:
            pass

    new_wb.save(r"D:\Durai\Scraping\TV\Save Data\Scraping Save\Tv Price List " + date +".xlsx")
end_time = time.time()
print('Duration: {}'.format(end_time - start_time))
driver.quit()






