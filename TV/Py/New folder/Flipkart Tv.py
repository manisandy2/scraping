from selenium import webdriver
from openpyxl import Workbook,load_workbook
import datetime

date = datetime.datetime.today().strftime("%d-%m-%Y")
date1 = datetime.datetime.today().strftime("%d-%m-%Y / %X")

driver = webdriver.Chrome(executable_path=r"C:\Durai\Driver\chromedriver.exe")

wb = load_workbook(r"C:\Durai\TV\Urls\Flipkart Tv Url.xlsx")
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active

new_ws.cell(row=1,column=1).value = date1
l = 2
for r in range(2,ws.max_row+1):
    if ws.cell(row=r,column=4).value != "N/A":
        print(r)
    try:
        driver.get(ws.cell(row=r,column=4).value)
        print(ws.cell(row=r, column=4).value)

        name = driver.find_element_by_class_name("B_NuCI")
        print(name.text)
        new_ws.cell(row=l,column=1).value = name.text

        price = driver.find_element_by_class_name("_30jeq3")
        print(price.text)
        new_ws.cell(row=l,column=2).value = price.text
        new_wb.save(r"C:\Durai\TV\Save Data\Flipkart Tv  "+date+".xlsx")
        l = l+1
    except:
        pass

