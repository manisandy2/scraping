from openpyxl import load_workbook
from selenium import webdriver
import time
import datetime


wb = load_workbook(r"C:\Durai\Mobile\Urls\Mobiles-URL.xlsx")

for ws in wb:
    if ws.title == 'Croma':
        print(ws.title)
        wb.active
        break
ws.cell(row=1, column=1).value = datetime.datetime.today()
r = 2

print(ws.max_row)
print(ws.max_column)
row = ws.max_row
col = ws.max_column

driver = webdriver.Chrome(executable_path=r"C:\Durai\Driver\chromedriver.exe")


for r in range(2, row + 1):
    url = ws.cell(row=r, column=2).value
    print(r)
    try:
        if not url == "NA":
            print(url)
            driver.get(url)
            time.sleep(1)
            name = driver.find_element_by_class_name("pd-title")
            print(name.text)
            ws.cell(row=r, column=3).value = name.text
            for price in driver.find_elements_by_class_name("main-product-price"):
                for price1 in price.find_elements_by_class_name('new-price'):
                    for price2 in price1.find_elements_by_class_name('amount'):
                        print(price2.text)
                        ws.cell(row=r, column=4).value = price2.text
            wb.save(r"C:\Durai\Mobile\Save Data\Croma-Mobiles-Data.xlsx")
    except:
        print('NA')
        ws.cell(row=r, column=6).value
        wb.save(r"C:\Durai\Mobile\Save Data\Croma-Mobiles-Data.xlsx")
driver.quit()