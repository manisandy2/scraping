from openpyxl import load_workbook
from selenium import webdriver
import time
import datetime


wb = load_workbook(r"C:\Durai\Mobile\Urls\Mobiles-URL.xlsx")

for ws in wb:
    if ws.title == 'Vijay sales':
        print(ws.title)
        wb.active
        break
ws.cell(row=1, column=1).value = datetime.datetime.today()

print(ws.max_row)
print(ws.max_column)
row = ws.max_row
col = ws.max_column
driver = webdriver.Chrome(executable_path=r"C:\Durai\Driver\chromedriver.exe")

s = 1

for r in range(2, ws.max_row + 1):
    url = ws.cell(row=r, column=2).value
    print(s)
    s = s + 1
    try:
        if not url == "NA":
            print(url)
            driver.get(url)
            # time.sleep(0.5)
            try:
                name = driver.find_element_by_tag_name("h1")
                print(name.text)
                ws.cell(row=r, column=3).value = name.text
                price1 = driver.find_element_by_class_name("priceMRP")
                for price2 in price1.find_elements_by_xpath(
                        '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span'):
                    print(price2.text)
                    ws.cell(row=r, column=4).value = price2.text
                wb.save(r"C:\Durai\Mobile\Save Data\Vijay-Sales-Mobiles-Data.xlsx")
            except:
                time.sleep(0.5)
                name1 = driver.find_element_by_class_name('pdp__title')
                print(name1.text)
                ws.cell(row=r, column=3).value = name1.text
                price2 = driver.find_element_by_class_name('pdp__offerPrice')
                print(price2.text)
                ws.cell(row=r, column=4).value = price2.text
                wb.save(r"C:\Durai\Mobile\Save Data\Vijay-Sales-Mobiles-Data.xlsx")
                time.sleep(0.5)
    except:
        print('NA')
driver.quit()


