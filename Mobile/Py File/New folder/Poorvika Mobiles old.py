
from openpyxl import Workbook
from selenium import webdriver
import datetime
import time
date = datetime.date.today().strftime("%d-%m-%Y")
l = 2

wb = Workbook()
ws = wb.active
ws.title = "MOBILE"

driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")

ws.cell(row=1, column=1).value = datetime.datetime.today()

driver.get(url="https://www.poorvika.com/mobile-phones")
length = len(driver.find_elements_by_class_name("paginationid"))

print(length)
for r in range(1, length+1):
    print(r)
    driver.get(url="https://www.poorvika.com/mobile-phones?&fmin=1&fmax=179900&order=l-h&page="+str(r))
    driver.implicitly_wait(1)
    # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    for cat in driver.find_elements_by_class_name("right-block"):
        name = cat.find_element_by_tag_name("h4")
        variant = cat.find_element_by_tag_name("h6")
        print(name.text ,variant.text)
        ws.cell(row=l, column=1).value = name.text + variant.text
        price = cat.find_element_by_class_name("cat-price-new")
        print(price.text[1:])
        ws.cell(row=l, column=2).value = price.text[1:]
        wb.save(r"C:\Durai\Mobile\Save Data\Poorvika Mobiles  "+date+".xlsx")
        l = l + 1
driver.quit()

