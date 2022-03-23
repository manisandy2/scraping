from openpyxl import load_workbook
from selenium import webdriver
import time
import datetime


wb = load_workbook(r"C:\Durai\Mobile\Urls\Mobiles-URL.xlsx")

for ws in wb:
    if ws.title == 'Reliance':
        print(ws.title)
        wb.active
        break
ws.cell(row=1, column=1).value = datetime.datetime.today()
r = 2

print(ws.max_row)
print(ws.max_column)
row = ws.max_row
col = ws.max_column
driver = webdriver.Chrome(executable_path=r"C:\Durai\Driver\chromedriver.exe")

s = 1
for r in range(2, ws.max_row + 1):
    url = ws.cell(row=r, column=2).value
    print(s)
    s = s + 1
    try:
        if not url == "NA":
            print(url)
            driver.get(url)
            time.sleep(0.5)
            name = driver.find_element_by_class_name("pdp__title")
            print(name.text)
            ws.cell(row=r, column=3).value = name.text
            price = driver.find_element_by_class_name("pdp__offerPrice")
            print(price.text)
            ws.cell(row=r, column=4).value = price.text
            wb.save(r"C:\Durai\Mobile\Save Data\Reliance-Mobiles-Data.xlsx")
    except:
        print('NA')
        ws.cell(row=r, column=4).value
        wb.save(r"C:\Durai\Mobile\Save Data\Reliance-Mobiles-Data.xlsx")
driver.quit()