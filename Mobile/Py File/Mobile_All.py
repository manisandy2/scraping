from openpyxl import load_workbook,Workbook
from selenium import webdriver
from selenium.webdriver import ChromeOptions
import datetime
import time
# 'cookies': 2, 'images': 2, 'javascript': 2, 'plugins': 2, 'popups': 2,
start_time = time.time()
options = webdriver.ChromeOptions()
prefs = {'profile.default_content_setting_values': {'geolocation': 2,
                            'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,
                            'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,

                            'media_stream_mic': 2, 'media_stream_camera': 2, 'protocol_handlers': 2,
                            'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                            'push_messaging': 2, 'ssl_cert_decisions': 2, 'metro_switch_to_desktop': 2,
                            'protected_media_identifier': 2, 'app_banner': 2, 'site_engagement': 2,
                            'durable_storage': 2}}
options.add_experimental_option('prefs', prefs)
# options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
# options.headless = True


date = datetime.date.today().strftime("%#d-%#m-%Y")
print(date)
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe",options=options)

wb = load_workbook(r"D:\Durai\Scraping\Mobile\Save Data\Save Urls\Scraping Urls "+date +".xlsx")
ws = wb.active
new_wb = Workbook()
new_ws = new_wb.active

new_ws.cell(row=1,column=1).value = "model Name"
new_ws.cell(row=1,column=2).value = "Poorvika price"
new_ws.cell(row=1,column=3).value = "Flipkart price"
new_ws.cell(row=1,column=4).value = "Amazon price"
new_ws.cell(row=1,column=5).value = "Croma price"
new_ws.cell(row=1,column=6).value = "vijay sale price"
new_ws.cell(row=1,column=7).value = "Reliance Digital price"

for r in range(1, ws.max_row+1):
    print(r)
    new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value
    new_ws.cell(row=r, column=2).value = ws.cell(row=r, column=2).value
#     if ws.cell(row=r,column=3).value != "N/A":
#
# # Flipkart
#         print(ws.cell(row=r,column=3).value)
#         try:
#             driver.get(url=ws.cell(row=r,column=3).value)
#             new_ws.cell(row=r,column=3).value = driver.find_element_by_class_name("_30jeq3").text[1:]
#             print(driver.find_element_by_class_name("_30jeq3").text)
#         except:
#             pass
#
# # Amazon
#
#     if ws.cell(row=r, column=4).value != "N/A":
#         print(ws.cell(row=r,column=4).value)
#
#         try:
#             driver.get(url=ws.cell(row=r,column=4).value)
#             price = driver.find_element_by_id("apex_desktop")
#             price1 = price.find_element_by_class_name("a-price-whole")
#             print(price1.text)
#             new_ws.cell(row=r, column=4).value = price1.text
#         except:
#             pass
#
#         try:
#             driver.get(url=ws.cell(row=r, column=4).value)
#             price = driver.find_element_by_id("apex_desktop")
#             price1 = price.find_element_by_class_name("a-size-medium")
#             print(price1.text[1:])
#             new_ws.cell(row=r, column=4).value = price1.text[1:]
#         except:
#             pass
#
# # Croma
#     if ws.cell(row=r, column=5).value != "N/A":
#         print(ws.cell(row=r, column=5).value)
#         try:
#             driver.get(url=ws.cell(row=r,column=5).value)
#             for price in driver.find_elements_by_class_name("cp-price"):
#                 for price1 in price.find_elements_by_class_name('new-price'):
#                     for price2 in price1.find_elements_by_class_name('amount'):
#                         new_ws.cell(row=r, column=5).value = price2.text
#                         print(price2.text)
#         except:
#             pass
# vijay sale

    if ws.cell(row=r, column=7).value != "N/A":
        print(ws.cell(row=r, column=7).value)
        try:
            driver.get(url=ws.cell(row=r, column=7).value)
            price1 = driver.find_element_by_class_name("priceMRP")
            for price2 in price1.find_elements_by_xpath('//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span'):
                for price2 in price1.find_elements_by_xpath( '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span'):
                    # for price2 in price1.find_elements_by_xpath('/html/body/form/div[4]/div[10]/div/div[4]/div/div[3]/div[1]/div[2]/div[3]/div/span[2]/span'):
                    print(price2.text)
                    new_ws.cell(row=r, column=6).value = price2.text
        except:
            pass

# Reliance

    if ws.cell(row=r, column=6).value != "N/A":
        print(ws.cell(row=r, column=6).value)
        try:
            driver.get(url=ws.cell(row=r, column=6).value)
            price = driver.find_element_by_class_name("pdp__offerPrice").text
            print(price[1:])
            new_ws.cell(row=r, column=7).value = price[1:]
        except:
            pass

    new_wb.save(r"D:\Durai\Scraping\Mobile\Save Data\Scraping save\Mobile Price List 10 " + date + ".xlsx")
end_time = time.time()
print('Duration: {}'.format(end_time - start_time))
driver.quit()
