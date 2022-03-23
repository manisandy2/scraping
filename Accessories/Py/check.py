from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import datetime
from selenium.webdriver.common.by import By
import time
date = datetime.date.today().strftime("%d-%m-%Y")

wb = Workbook()
ws = wb.active
# ws.title = "personal-health-care"

service = Service(r"D:\Durai\Driver\chromedriver.exe")

driver = webdriver.Chrome(service=service)
ws.cell(row=1, column=1).value = datetime.datetime.today()
url="https://www.poorvika.com/mobiles-accessories?&fmin=2&fmax=179900&stock=1&order=l-h&page="
driver.get(url)
ws.cell(row=1, column=1).value = datetime.datetime.today()
l = 2
length = len(driver.find_elements(By.CLASS_NAME, "paginationid"))
driver.implicitly_wait(3)
last_value = length + 1
for r in range(1, length + 1):
    print(r)
    driver.get(url + str(r))
    driver.implicitly_wait(1)


    for r in driver.find_elements(By.CLASS_NAME,"right-block"):
        name = r.find_element_by_tag_name("h4")
        variant = r.find_element_by_tag_name("h6")
        price = r.find_element_by_class_name("cat-price-new").text[1:]
        print(name.text,variant.text)
        print(price)
        ws.cell(row=l, column=1).value = name.text + variant.text
        ws.cell(row=l, column=2).value = price
        l = l + 1
        wb.save(r"D:\Durai\Scraping\Accessories\Save File\Poorvika File\check " + date + ".xlsx")

driver.quit()


