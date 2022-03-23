from selenium.webdriver.common.by import By
import datetime
from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
# ws.title = "All ACCESSORIES"

date = datetime.date.today().strftime("%d-%m-%Y")
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")


class Model:
    def __init__(self,title, url, url_list, row_num,wb_title):
        self.url = url
        self.url_list = url_list
        self.row_num = row_num
        self.title = title
        self.wb_title = wb_title

    def title_price(self):
        self.worksheet_name()
        for cat in driver.find_elements(By.CLASS_NAME, "right-block"):
            name = cat.find_element(By.TAG_NAME, "h4")
            variant = cat.find_element(By.TAG_NAME, "h6")
            print(name.text, variant.text)
            ws.cell(row=self.row_num, column=1).value = name.text + variant.text
            price = cat.find_element(By.CLASS_NAME, "cat-price-new")
            print(price.text[1:])
            ws.cell(row=self.row_num, column=2).value = price.text[1:]
            ws.cell(row=self.row_num, column=3).value = self.title
            self.product_save()
            # wb.save(r"D:\Durai\Scraping\Accessories\Save File\Poorvika File\Accessories " + date + ".xlsx")
            self.row_num = self.row_num + 1

    def product(self):
        # accessories-brands , mobile-accessories , smart-technology,audio
        driver.get(url=self.url)
        length = len(driver.find_elements(By.CLASS_NAME, "paginationid"))
        driver.implicitly_wait(3)
        last_value = length + 1
        for r in range(1, length + 1):
            print(r)
            driver.get(url=self.url_list + str(r))
            driver.implicitly_wait(1)
            self.title_price()
        print(self.title + " Complete")

    def product1(self):
        # personal health care

        driver.get(url=self.url)
        self.title_price()
        print(self.title + " Complete")

    def product_num(self):
        return self.row_num

    def product_save(self):
        if self.wb_title == "All Accessories ":
            wb.save(r"D:\Durai\Scraping\Accessories\Save File\Poorvika File\ " + self.wb_title[4:] + date + ".xlsx")
        elif self.wb_title == "All Laptops ":
            wb.save(r"D:\Durai\Scraping\Laptops\Save File\Poorvika\ " + self.wb_title[4:] + date + ".xlsx")

    def worksheet_name(self):
        ws.title = self.wb_title


