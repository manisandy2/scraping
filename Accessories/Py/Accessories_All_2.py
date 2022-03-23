from openpyxl import load_workbook,Workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ChromeOptions
from selenium.webdriver.common.by import By
from Scraping.Accessories.Py.competitor import PriceCompression
import datetime
import time
# 'cookies': 2, 'images': 2, 'javascript': 2, 'plugins': 2, 'popups': 2,
start_time = time.time()
options = webdriver.ChromeOptions()
prefs = {'profile.default_content_setting_values': {'geolocation': 2,
                            'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2,
                            'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,
                            'media_stream_mic': 2, 'media_stream_camera': 2, 'protocol_handlers': 2,
                            'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2,
                            'push_messaging': 2, 'ssl_cert_decisions': 2, 'metro_switch_to_desktop': 2,
                            'protected_media_identifier': 2, 'app_banner': 2, 'site_engagement': 2,
                            'durable_storage': 2}}
options.add_experimental_option('prefs', prefs)
# options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
# options.headless = True


date = datetime.date.today().strftime("%#d-%#m-%Y")
print(date)
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe",chrome_options=options)

wb = load_workbook(r"D:\Durai\Scraping\Accessories\Scraping Urls\Accessories Urls "+date +".xlsx")
ws = wb.active
new_wb = Workbook()
new_ws = new_wb.active


new_ws.cell(row=1,column=1).value = "Model Name"
new_ws.cell(row=1,column=2).value = "Poorvika Price"
new_ws.cell(row=1,column=3).value = "Flipkart Name"
new_ws.cell(row=1,column=4).value = "Flipkart Price"
new_ws.cell(row=1,column=5).value = "Flipkart Url"
new_ws.cell(row=1,column=6).value = "Amazon Name"
new_ws.cell(row=1,column=7).value = "Amazon Price"
new_ws.cell(row=1,column=8).value = "Amazon Url"
new_ws.cell(row=1,column=9).value = "Croma Name"
new_ws.cell(row=1,column=10).value = "Croma Price"
new_ws.cell(row=1,column=11).value = "Croma Url"
new_ws.cell(row=1,column=12).value = "Vijay Sale Name"
new_ws.cell(row=1,column=13).value = "Vijay Sale price"
new_ws.cell(row=1,column=14).value = "Vijay Sale Url"
new_ws.cell(row=1,column=15).value = "Reliance Digital Name"
new_ws.cell(row=1,column=16).value = "Reliance Digital Price"
new_ws.cell(row=1,column=17).value = "Reliance Digital Url"

for r in range(501, ws.max_row+1):
    print(r)
    new_ws.cell(row=r, column=1).value = ws.cell(row=r, column=1).value
    new_ws.cell(row=r, column=2).value = ws.cell(row=r, column=2).value
    pc = PriceCompression(driver=driver,row_num=r,new_ws=new_ws,ws=ws)
    pc.flipkart()
    pc.amazon()
    pc.croma()
    pc.vijay_sale()
    pc.reliance()
    new_wb.save(r"D:\Durai\Scraping\Accessories\Save File\Scraping Data\Scraping File\Accessories all 2 Price List " + date + ".xlsx")
driver.quit()
