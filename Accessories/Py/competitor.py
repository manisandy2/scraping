from openpyxl import Workbook,load_workbook
from selenium.webdriver.common.by import By
import datetime
import time


class PriceCompression:
    def __init__(self,driver,row_num,new_ws,ws):
        self.driver = driver
        self.row_num = row_num
        self.new_ws = new_ws
        self.ws = ws

    def flipkart(self):
        if self.ws.cell(row=self.row_num, column=3).value != "N/A":
            print(self.ws.cell(row=self.row_num, column=3).value)
            self.new_ws.cell(row=self.row_num, column=5).value = self.ws.cell(row=self.row_num, column=3).value
            try:
                time.sleep(3)
                self.driver.implicitly_wait(1)
                self.driver.get(url=self.ws.cell(row=self.row_num, column=3).value)
                self.driver.implicitly_wait(3)
                self.new_ws.cell(row=self.row_num, column=3).value = self.driver.find_element(By.TAG_NAME,"h1").text
                self.new_ws.cell(row=self.row_num, column=4).value = self.driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:]
                print("flipkart")
                print(self.driver.find_element(By.TAG_NAME,"h1").text)
                print(self.driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:])
            except:
                pass

    def amazon(self):
        if self.ws.cell(row=self.row_num, column=4).value != "N/A":
            print(self.ws.cell(row=self.row_num, column=4).value)
            self.new_ws.cell(row=self.row_num, column=8).value = self.ws.cell(row=self.row_num, column=4).value
            try:
                self.driver.get(url=self.ws.cell(row=self.row_num, column=4).value)
                self.driver.implicitly_wait(3)
                print(self.driver.find_element(By.ID, "productTitle").text)
                self.new_ws.cell(row=self.row_num, column=6).value = self.driver.find_element(By.ID, "productTitle").text
                try:
                    for box_price in self.driver.find_elements(By.ID, "apex_desktop"):
                        for price in box_price.find_elements(By.CLASS_NAME,"a-price"):
                            print("Amazon 1")
                            print(price.find_element(By.CLASS_NAME,"a-price-whole").text)
                            self.new_ws.cell(row=self.row_num, column=7).value = price.find_element(By.CLASS_NAME,"a-price-whole").text
                except:
                    print("Amazon 2")
                    print(self.driver.find_element(By.CLASS_NAME, "apexPriceToPay").text)
                    self.new_ws.cell(row=self.row_num, column=7).value = self.driver.find_element(By.CLASS_NAME, "apexPriceToPay").text[1:]

            except :
                pass

    def croma(self):
        if self.ws.cell(row=self.row_num, column=5).value != "N/A":
            print(self.ws.cell(row=self.row_num, column=5).value)
            self.new_ws.cell(row=self.row_num, column=11).value = self.ws.cell(row=self.row_num, column=5).value
            try:
                self.driver.get(url=self.ws.cell(row=self.row_num,column=5).value)
                self.driver.implicitly_wait(3)
                print("croma")
                print(self.driver.find_element(By.TAG_NAME,"h1").text)
                self.new_ws.cell(row=self.row_num, column=9).value = self.driver.find_element(By.TAG_NAME,"h1").text

                for price in self.driver.find_elements(By.CLASS_NAME, "cp-price"):
                    for price1 in price.find_elements(By.CLASS_NAME, 'new-price'):
                        for price2 in price1.find_elements(By.CLASS_NAME, 'amount'):
                            print(price2.text[1:])
                            self.new_ws.cell(row=self.row_num, column=10).value = price2.text[1:]

            except:
                pass

    def vijay_sale(self):
        if self.ws.cell(row=self.row_num, column=6).value != "N/A":
            print(self.ws.cell(row=self.row_num, column=6).value)
            self.new_ws.cell(row=self.row_num, column=14).value = self.ws.cell(row=self.row_num, column=6).value
            try:
                self.driver.get(url=self.ws.cell(row=self.row_num, column=6).value)
                self.driver.implicitly_wait(3)
                print(self.driver.find_element(By.TAG_NAME,"h1").text)
                self.new_ws.cell(row=self.row_num, column=12).value = self.driver.find_element(By.TAG_NAME,"h1").text
                try:
                    if self.driver.find_element(By.ID,"ContentPlaceHolder1_fillprice").text != None:
                        # print("Nothing")
                        try:
                            print("vijay sale")
                            print(self.driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text)
                            price = self.driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div[1]/div[1]/span[2]/span').text
                            self.new_ws.cell(row=self.row_num, column=13).value = price

                        except:
                            print("vijay sale")
                            print(self.driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text)
                            price = self.driver.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_fillprice"]/div/span[2]/span').text
                            self.new_ws.cell(row=self.row_num, column=13).value = price

                except:
                    pass
            except:
                pass

    def reliance(self):
        if self.ws.cell(row=self.row_num, column=7).value != "N/A":
            print(self.ws.cell(row=self.row_num, column=7).value)
            self.new_ws.cell(row=self.row_num, column=17).value = self.ws.cell(row=self.row_num, column=7).value
            try:
                self.driver.get(url=self.ws.cell(row=self.row_num, column=7).value)
                self.driver.implicitly_wait(3)
                print(self.driver.find_element(By.TAG_NAME,'h1').text)
                self.new_ws.cell(row=self.row_num, column=15).value = self.driver.find_element(By.TAG_NAME,'h1').text
                price = self.driver.find_element(By.CLASS_NAME, "pdp__offerPrice").text
                self.new_ws.cell(row=self.row_num, column=16).value = price[1:]
                print("reliance")
                print(price[1:])
            except:
                pass