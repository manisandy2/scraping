from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import datetime
from selenium.webdriver.common.by import By
import time
date = datetime.date.today().strftime("%d-%m-%Y")

wb = Workbook()
ws = wb.active
ws.title = "personal-health-care"

service = Service(r"D:\Durai\Driver\chromedriver.exe")

driver = webdriver.Chrome(service=service)
ws.cell(row=1, column=1).value = datetime.datetime.today()

driver.get(url="https://www.poorvika.com/personal-health-care")
ws.cell(row=1, column=1).value = datetime.datetime.today()

product_list = driver.find_elements(By.CLASS_NAME,"right-block")
# product_list = driver.find_elements_by_class_name("")
# print(len(product_list))
print(len(product_list))

l = 2


for r in driver.find_elements(By.CLASS_NAME,"right-block"):
    name = r.find_element_by_tag_name("h4")
    variant = r.find_element_by_tag_name("h6")
    price = r.find_element_by_class_name("cat-price-new").text[1:]
    print(name.text,variant.text)
    print(price)
    ws.cell(row=l, column=1).value = name.text + variant.text
    ws.cell(row=l, column=2).value = price
    wb.save(r"D:\Durai\Accessories\Save Data\personal-health-care " + date + ".xlsx")
    l = l +1
driver.quit()


