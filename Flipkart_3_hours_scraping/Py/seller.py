from selenium import webdriver
import datetime
from openpyxl import load_workbook,Workbook
from selenium.webdriver.common.by import By
import time
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"D:\Durai\Scraping\Flipkart_3_hours_scraping\Urls\Flipkart Scraping.xlsx")
ws = wb.active
date = datetime.datetime.now().strftime("%d-%m-%Y %I %p")

new_wb = Workbook()
new_ws = new_wb.active


class Filpkart_Scraping_3_hours:

    def heading(self):
        new_ws.cell(row=1, column=1).value = "Name"
        new_ws.cell(row=1, column=2).value = "Item Code"
        new_ws.cell(row=1, column=3).value = "Flipkart Urls"
        new_ws.cell(row=1, column=4).value = "Flipkart Name"
        new_ws.cell(row=1, column=5).value = "Flipkart Price"

    def file_save(self,**kwargs):
        l = 7
        for r1 in driver.find_elements(By.CLASS_NAME, "_2Y3EWJ"):
            new_ws.cell(row=kwargs.get('row_range'), column=l).value = r1.find_element(By.CLASS_NAME, "_3enH42").text
            l = l + 1
            new_ws.cell(row=kwargs.get('row_range'), column=l).value = r1.find_element(By.CLASS_NAME, "_30jeq3").text
            l = l + 1
            new_wb.save(r"D:\Durai\Scraping\Flipkart_3_hours_scraping\save data\Flipkart_All_Sellers " + str(kwargs.get('path')) + " " + date + ".xlsx")

    def before_excel(self,**kwargs):
        new_ws.cell(row=kwargs.get('row_range'), column=1).value = ws.cell(row=kwargs.get('row_range'), column=1).value
        new_ws.cell(row=kwargs.get('row_range'), column=2).value = ws.cell(row=kwargs.get('row_range'), column=2).value
        new_ws.cell(row=kwargs.get('row_range'), column=3).value = ws.cell(row=kwargs.get('row_range'), column=3).value

    def excel_value(self,**kwargs):

        new_ws.cell(row=kwargs.get('row_range'), column=4).value = driver.find_element(By.CLASS_NAME, "B_NuCI").text
        new_ws.cell(row=kwargs.get('row_range'), column=5).value = driver.find_element(By.CLASS_NAME, "_30jeq3").text[1:]
        driver.find_element(By.CLASS_NAME, "_36yFo0").send_keys(600032)
        time.sleep(2)
        driver.find_element(By.CLASS_NAME, "_2P_LDn").click()
        time.sleep(2)
        driver.find_element(By.CLASS_NAME, "_1_xoMS").click()
        time.sleep(2)

    def flikart_seller(self,**kwargs):
        Filpkart_Scraping_3_hours.heading(self)

        for r in range(kwargs.get('start'), kwargs.get('end')):
            print(r)
            Filpkart_Scraping_3_hours.before_excel(self,row_range=r)
            try:
                driver.get(url=ws.cell(row=r, column=3).value)
                Filpkart_Scraping_3_hours.excel_value(self,row_range=r)
                Filpkart_Scraping_3_hours.file_save(self, row_range=r, path=kwargs.get('path'))

            except:
                pass
        driver.close()
        driver.quit()

