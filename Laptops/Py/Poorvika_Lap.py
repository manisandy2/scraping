from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import  datetime

wb = Workbook()
ws = wb.active
ws.title = "Poorvika Laptop"

l = 2

date = datetime.date.today().strftime("%d-%m-%Y")
ws.cell(row=1,column=1).value = datetime.datetime.now()
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")


class Model:
    def __init__(self,url):

        self.url = url



    def Laptop_len(self):

        driver.get(url=self.url)
        Length = len(driver.find_elements(By.CLASS_NAME, "paginationid"))
        print(Length)
        # for r in range(1, Length+ 1):
        #     print(r)
        #     driver.get(url=self.urls + str(r))
        #     driver.implicitly_wait(1)
    #
    # def Laptop(self):
    #l
    #     for cat in driver.find_elements(By.CLASS_NAME, "right-block"):
    #         name = cat.find_element(By.TAG_NAME, "h4")
    #         variant = cat.find_element(By.TAG_NAME, "h6")
    #         print(name.text, variant.text)
    #         ws.cell(row=l, column=1).value = name.text + variant.text
    #         price = cat.find_element(By.CLASS_NAME, "cat-price-new")
    #         print(price.text[1:])
    #         ws.cell(row=l, column=2).value = price.text[1:]
    #         wb.save(r"D:\Durai\Laptops\Save Data\poorvika_Laptop " + date + ".xlsx")
    #         l = l + 1
    #

